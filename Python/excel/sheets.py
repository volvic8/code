#! /usr/bin/env python3
"""ディレクトリ内のexcelファイルを読み込み、シート一覧を出力

Args:
    parent_path(str): 読み込み先のディレクトリ（未指定の場合はカレントディレクトリ）
    output_file(str): 出力するexcelファイル名を指定
"""

import os
import sys
import openpyxl

XLS_EXTS = (".xls", ".xlsx", ".xlsm", ".xlsb",
            ".xlt", ".xltx", ".xltm", ".xla", ".xlam")


def join_sheets(full_path):
    """シート名を改行文字で結合する

    Args:
        full_path(str): excelファイルのフルパス

    Returns:
        sheets_join(str): シート名を改行文字で結合した文字列
    """
    book = openpyxl.load_workbook(full_path)
    sheets = book.sheetnames
    sheets_join = ""
    for i in range(len(sheets)):
        # 非表示シートを除外
        if book[sheets[i]].sheet_state != "hidden":
            if i == 0:
                sheets_join += sheets[i]
            else:
                sheets_join += "\r\n" + sheets[i]

    return sheets_join


def output_excel(files):
    """ファイル情報をexcelファイルに出力する。

    Args:
        files(list): ファイル情報（ファイル名, 結合したシート名）のリスト
    """
    book = openpyxl.Workbook()
    sheet = book.active
    sheet.title = "Sheet1"
    for i in range(len(files)):
        sheet.cell(row=i+1, column=1).value = files[i][0]
        sheet.cell(row=i+1, column=2).value = files[i][1]

    book.save(output_file)


if __name__ == '__main__':
    if len(sys.argv) > 1:
        parent_path = sys.argv[1]
    else:
        # ディレクトリの指定がなければカレントディレクトリを設定
        parent_path = os.getcwd()

    if len(sys.argv) > 2:
        output_file = sys.argv[2]
    else:
        output_file = "out.xlsx"

    files = []
    # 指定ディレクトリ配下のファイルを処理
    for folder_path, _, sub_file in os.walk(parent_path):
        for file in sub_file:
            # excel以外のファイルはスキップ
            ext = os.path.splitext(file)[1].lower()
            if ext not in XLS_EXTS:
                continue

            files.append([file, join_sheets(folder_path + "\\" + file)])

    output_excel(files)

# TODO exe
